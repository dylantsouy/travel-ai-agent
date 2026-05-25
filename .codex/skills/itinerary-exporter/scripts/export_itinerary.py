#!/usr/bin/env python3
"""Export normalized travel itinerary rows to CSV and/or XLSX.

Input can be JSON, CSV, or Markdown tables. JSON is preferred:

[
  {
    "date": "2026/7/10",
    "location": "Hotel",
    "event": "check in",
    "arrival": "15:00",
    "departure": "16:00"
  }
]

Also accepts {"rows": [...]} or {"days": [{"date": "...", "rows": [...]}]}.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any


ITINERARY_FIELDS = [
    "date",
    "day",
    "location",
    "event",
    "arrival",
    "duration",
    "departure",
    "transport",
    "notes",
    "alternatives",
    "link",
    "status",
]

MAP_FIELDS = ["date", "sequence", "place_name", "search_query", "category", "notes"]
CALENDAR_FIELDS = [
    "title",
    "start_date",
    "start_time",
    "end_date",
    "end_time",
    "location",
    "description",
]

FIELD_SETS = {
    "itinerary": ITINERARY_FIELDS,
    "maps": MAP_FIELDS,
    "calendar": CALENDAR_FIELDS,
}

KEY_ALIASES = {
    "日": "date",
    "日期": "date",
    "星期": "day",
    "地點": "location",
    "地点": "location",
    "待處理事件": "event",
    "待处理事件": "event",
    "事件": "event",
    "抵達": "arrival",
    "抵达": "arrival",
    "停留": "duration",
    "出發": "departure",
    "出发": "departure",
    "採用交通": "transport",
    "采用交通": "transport",
    "交通": "transport",
    "備註": "notes",
    "备注": "notes",
    "備選": "alternatives",
    "备选": "alternatives",
    "連結": "link",
    "链接": "link",
    "狀態": "status",
    "状态": "status",
    "地點名稱": "place_name",
    "地点名称": "place_name",
    "搜尋關鍵字": "search_query",
    "搜索关键字": "search_query",
    "分類": "category",
    "分类": "category",
    "標題": "title",
    "标题": "title",
    "開始日期": "start_date",
    "开始日期": "start_date",
    "開始時間": "start_time",
    "开始时间": "start_time",
    "結束日期": "end_date",
    "结束日期": "end_date",
    "結束時間": "end_time",
    "结束时间": "end_time",
    "描述": "description",
}


def plain_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(
        r"[\U0001F1E6-\U0001FAFF\u2600-\u27BF]+",
        "",
        text,
        flags=re.UNICODE,
    )
    return re.sub(r"\s+", " ", text).strip()


def split_markdown_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [cell.strip() for cell in line.split("|")]


def is_separator(line: str) -> bool:
    cells = split_markdown_row(line)
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in cells)


def normalize_row(row: dict[str, Any], extras: dict[str, Any] | None = None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        canonical = KEY_ALIASES.get(str(key).strip(), str(key).strip())
        if canonical in {"location", "place_name"} and "link" not in normalized:
            match = re.search(r"\[[^\]]+\]\(([^)]+)\)", "" if value is None else str(value))
            if match:
                normalized["link"] = match.group(1).strip()
        normalized[canonical] = plain_text(value)
    if extras:
        for key, value in extras.items():
            normalized.setdefault(key, plain_text(value))
    return normalized


def flatten_json(data: Any) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if isinstance(data, list):
        return [normalize_row(row) for row in data if isinstance(row, dict)]
    if not isinstance(data, dict):
        raise ValueError("JSON input must be a list, an object with rows, or an object with days.")
    if isinstance(data.get("rows"), list):
        return [normalize_row(row) for row in data["rows"] if isinstance(row, dict)]
    if isinstance(data.get("days"), list):
        for day in data["days"]:
            if not isinstance(day, dict):
                continue
            extras = {
                "date": day.get("date") or day.get("title") or "",
                "day": day.get("day") or "",
            }
            for row in day.get("rows", []):
                if isinstance(row, dict):
                    rows.append(normalize_row(row, extras))
        return rows
    raise ValueError("JSON object must contain rows or days.")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return [normalize_row(row) for row in csv.DictReader(file)]


def read_json(path: Path) -> list[dict[str, str]]:
    return flatten_json(json.loads(path.read_text(encoding="utf-8-sig")))


def read_markdown(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    rows: list[dict[str, str]] = []
    current_date = ""
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("###"):
            current_date = line.lstrip("#").strip()
            i += 1
            continue
        if line.startswith("|") and i + 1 < len(lines) and is_separator(lines[i + 1]):
            headers = split_markdown_row(line)
            i += 2
            while i < len(lines) and lines[i].strip().startswith("|"):
                cells = split_markdown_row(lines[i])
                row = dict(zip(headers, cells))
                extras = {"date": current_date} if current_date else None
                rows.append(normalize_row(row, extras))
                i += 1
            continue
        i += 1
    return rows


def load_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return read_json(path)
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".md", ".markdown", ".txt"}:
        return read_markdown(path)
    raise ValueError("Unsupported input format. Use .json, .csv, .md, or .txt.")


def rows_for_kind(rows: list[dict[str, str]], kind: str) -> list[dict[str, str]]:
    if kind == "itinerary":
        return rows
    if kind == "maps":
        counters: dict[str, int] = {}
        mapped = []
        for row in rows:
            date = row.get("date", "")
            counters[date] = counters.get(date, 0) + 1
            place = row.get("place_name") or row.get("location", "")
            mapped.append(
                {
                    "date": date,
                    "sequence": str(counters[date]),
                    "place_name": place,
                    "search_query": row.get("search_query") or place,
                    "category": row.get("category", ""),
                    "notes": row.get("notes", ""),
                }
            )
        return mapped
    if kind == "calendar":
        mapped = []
        for row in rows:
            date = row.get("date", "")
            mapped.append(
                {
                    "title": row.get("title") or row.get("event") or row.get("location", ""),
                    "start_date": row.get("start_date") or date,
                    "start_time": row.get("start_time") or row.get("arrival", ""),
                    "end_date": row.get("end_date") or date,
                    "end_time": row.get("end_time") or row.get("departure", ""),
                    "location": row.get("location", ""),
                    "description": row.get("description") or row.get("notes", ""),
                }
            )
        return mapped
    raise ValueError(f"Unsupported kind: {kind}")


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_xlsx(path: Path, rows: list[dict[str, str]], fields: list[str], sheet_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export. Export CSV instead.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "Itinerary"
    worksheet.append(fields)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fields])
    for index, field in enumerate(fields, start=1):
        max_len = max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows])
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 10), 42)
    worksheet.freeze_panes = "A2"
    workbook.save(path)


def output_paths(input_path: Path, output: str | None, output_dir: str | None, fmt: str) -> dict[str, Path]:
    if output:
        out = Path(output)
        if fmt == "both":
            stem = out.with_suffix("")
            return {"csv": stem.with_suffix(".csv"), "xlsx": stem.with_suffix(".xlsx")}
        return {fmt: out.with_suffix(f".{fmt}")}
    directory = Path(output_dir) if output_dir else input_path.parent
    stem = input_path.stem
    if fmt == "both":
        return {"csv": directory / f"{stem}.csv", "xlsx": directory / f"{stem}.xlsx"}
    return {fmt: directory / f"{stem}.{fmt}"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export travel itinerary rows.")
    parser.add_argument("--input", required=True, help="Input .json, .csv, .md, or .txt file.")
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="csv")
    parser.add_argument("--kind", choices=sorted(FIELD_SETS), default="itinerary")
    parser.add_argument("--output", help="Output file path. For --format both, extension is ignored.")
    parser.add_argument("--output-dir", help="Output directory when --output is omitted.")
    parser.add_argument("--sheet-name", default="Itinerary")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    rows = rows_for_kind(load_rows(input_path), args.kind)
    fields = FIELD_SETS[args.kind]
    paths = output_paths(input_path, args.output, args.output_dir, args.format)
    written: dict[str, str] = {}

    if "csv" in paths:
        write_csv(paths["csv"], rows, fields)
        written["csv"] = str(paths["csv"].resolve())
    if "xlsx" in paths:
        write_xlsx(paths["xlsx"], rows, fields, args.sheet_name)
        written["xlsx"] = str(paths["xlsx"].resolve())

    print(json.dumps({"rows": len(rows), "written": written}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
